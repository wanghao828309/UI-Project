# coding=utf8
from openpyxl import Workbook
import openpyxl

class ExcelInput():
    index1 = 1
    index2 = 1
    index3 = 1
    index4 = 1
    index5 = 1
    index6 = 1

    def __init__(self):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "MUSIC"
        self.wb.create_sheet("TEXT")
        self.wb.create_sheet("TRANSITIONS")
        self.wb.create_sheet("FILTERS")
        self.wb.create_sheet("OVERLAYS")
        self.wb.create_sheet("ELEMENTS")

    def insertsheet(self,sheet_index,file_path,error):
        global index1
        global index2
        global index3
        global index4
        global index5
        global index6
        try:
            ws = self.wb.worksheets[sheet_index]
            if sheet_index == 0:

                ws["A%d"%self.index1] = file_path
                ws["B%d"%self.index1] = error
                self.index1 += 1

            if sheet_index == 1:
                ws["A%d" % self.index2] = file_path
                ws["B%d" % self.index2] = error
                self.index2 += 1

            if sheet_index == 2:
                ws["A%d" % self.index3] = file_path
                ws["B%d" % self.index3] = error
                self.index3 += 1

            if sheet_index == 3:
                ws["A%d" % self.index4] = file_path
                ws["B%d" % self.index4] = error
                self.index4 += 1

            if sheet_index == 4:
                ws["A%d" % self.index5] = file_path
                ws["B%d" % self.index5] = error
                self.index5 += 1

            if sheet_index == 5:
                ws["A%d" % self.index6] = file_path
                ws["B%d" % self.index6] = error
                self.index6 += 1

        except openpyxl.utils.exceptions.IllegalCharacterError:
            print file_path

    def save(self,path,filename):
        self.wb.save(path + "\\%s.xlsx" % filename)





   


