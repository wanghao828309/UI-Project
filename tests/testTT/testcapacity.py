# coding= utf-8
from __future__ import division
import smtplib
import subprocess
from openpyxl import Workbook
import openpyxl
import wmi
import os
import re
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
import socket
def get_mac_address():
    import uuid
    node = uuid.getnode()
    aaa = uuid.UUID(int = node).hex
    mac = uuid.UUID(int = node).hex[-12:]

    b = re.findall(r'.{2}', mac)
    c = '-'.join(b)
    return c

cwmi = wmi.WMI()
sys_ver = ""
def sys_version():
    # 获取操作系统版本
    global sys_ver
    for sys in cwmi.Win32_OperatingSystem():
        sys_ver = sys.Caption.encode("UTF") + str(sys.BuildNumber) + "," + sys.OSArchitecture.encode("UTF")

cpu_ram = 0
cpu_name = ""
def cpu_mem():
    # CPU类型和内存
    global cpu_ram
    global cpu_name
    for processor in cwmi.Win32_Processor():
        cpu_name = processor.Name.strip()
    for cobj in cwmi.Win32_ComputerSystem():
        # cpu_ram =  cobj.TotalPhysicalMemory.strip()
        sss = int(cobj.TotalPhysicalMemory.strip()) / 1024 / 1024 / 1024
        cpu_ram = str(sss) + "G"
    # for Memory in c.Win32_PhysicalMemory():
    #     print "Memory Capacity: %.fMB" % (int(Memory.Capacity) / 1048576)
gpu_ram = 0
gpu_name = ''
def gpu_mem():
    global gpu_ram
    global gpu_name
    # GPU类型和内存
    for gpu in cwmi.Win32_VideoController():
        gpu_name =  gpu.Name.strip()
        # ffff = gpu.AdapterRAM
        gpu_ram =  abs(gpu.AdapterRAM)/1024/1024
        gpu_ram = str(gpu_ram) + "M"

def GetFileList(dir, fileList):
    newDir = dir
    if os.path.isfile(dir):         # 如果是文件则添加进 fileList
        fileList.append(dir)
    elif os.path.isdir(dir):
        for s in os.listdir(dir):   # 如果是文件夹
            newDir = os.path.join(dir, s)
            GetFileList(newDir, fileList)
    return fileList

def result_has_error(result_string):
    if not result_string:
        return True

    all_result = re.findall(r'\[.*?(.+?)\]', result_string)
    for result in all_result:
        if (result.upper().strip() == "FAILED"):
            return True
    return False

sys_version()
cpu_mem()
gpu_mem()

all_test=[["CPU解码耗时",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
            ["GPU解码耗时(DX9)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
            ["GPU->CPU转YV12(DX9)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU解码耗时(DX10)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
            ["GPU->CPU转YV12(DX10)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU解码耗时(DX11)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
            ["GPU->CPU转YV12(DX11)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["CPU内缩放耗时（1/4采样）",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU内缩放耗时（1/4采样）",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["CPU内YV12转RGB耗时",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU内NV12转RGB耗时",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["CPU内RGB转YV12耗时",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU内RGB转NV12耗时",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["CPU到GPU内存拷贝耗时(OPENCL)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}],
          ["GPU到CPU内存拷贝耗时(OPENCL)",{"3840*2160":[],"2560*1440":[],"1920*1080":[],"1280*720":[],"960*540":[]}]]
wb = Workbook()
ws = wb.active
hostname = socket.gethostname()

ws["A3"]= "计算机名称"
ws["B3"]= hostname
ws["A4"]= "MAC地址"
ws["B4"]= get_mac_address()
ws["A5"]= "CPU型号"
ws["B5"]= cpu_name
ws["A6"]= "GPU型号"
ws["B6"]= gpu_name
ws["A7"]= "显存大小"
ws["B7"]= gpu_ram
ws["A8"]= "内存大小"
ws["B8"]= cpu_ram
ws["A9"]= "操作系统版本号"
ws["B9"]= sys_ver

# cpu 版 bgra to yv12_nv12
file=open('out.txt','w')
try:
    result = subprocess.Popen(["NLEProTestTool.exe","-t", str(39)],stdout=subprocess.PIPE).communicate()[0]
    print >> file, result
    print result
except subprocess.CalledProcessError as e:
    print e
file.close()
#
# # gpu 版 bgra to yv12_nv12
file=open('out.txt','a')
result = subprocess.check_output(["NLEProTestTool.exe","-t", str(39), "--gpu"])
print >> file, result
print result
file.close()

list = GetFileList('QA', [])
file=open('out.txt','a')

# yv12_nv12 to rgb
for testfil in  list:
    try:
        # result = subprocess.check_output(["NLEProTestTool.exe",testfil,"-t", str(37)])
        result = subprocess.Popen(["NLEProTestTool.exe",testfil,"-t", str(37)],stdout=subprocess.PIPE).communicate()[0]
        print >> file, result
        print result
    except subprocess.CalledProcessError as e:
        print e
for testfil in  list:
        result = subprocess.check_output(["NLEProTestTool.exe",testfil,"-t", str(37), "--gpu"])
        print >> file, result
        print result

file.close()

# gpu 版 拷贝
file=open('out.txt','a')
result = subprocess.check_output(["NLEProTestTool.exe","-t", str(38), "--gpu"])
print >> file, result
print result
file.close()

# cpu 版 拷贝
file=open('out.txt','a')
result = subprocess.check_output(["NLEProTestTool.exe","-t", str(38)])
print >> file, result
print result
file.close()

# gpu 版resize
file=open('out.txt','a')
result = subprocess.check_output(["NLEProTestTool.exe","-t", str(40), "--gpu"])
print >> file, result
print result
file.close()

# cpu 版resize
file=open('out.txt','a')
result = subprocess.check_output(["NLEProTestTool.exe","-t", str(40)])
print >> file, result
print result
file.close()

newf = open("out.txt")
# setTestItem=0
# times=0
for line in newf:
    if -1 != line.find("cpu bgra to yv12_nv12"):
        for key in all_test[11][1]:
            if -1 != line.find(key):
                all_test[11][1][key].append(str(line[line.find("spends:")+7:]))
                break
    if -1 != line.find("gpu bgra to yv12_nv12"):
        for key in all_test[12][1]:
            if -1 != line.find(key):
                all_test[12][1][key].append(str(line[line.find("spends:")+7:]))
                break

    if -1 != line.find("GPU yv12_nv12_to_argb with"):
        for key in all_test[10][1]:
            if -1 != line.find(key):
                all_test[10][1][key].append(str(line[line.find("spends:") + 7:]))
    if -1 != line.find("CPU yv12_nv12_to_argb with"):
        for key in all_test[9][1]:
            if -1 != line.find(key):
                all_test[9][1][key].append(str(line[line.find("spends:") + 7:]))

    if -1 != line.find("gpu 1/4 resize with"):
        for key in all_test[8][1]:
            if -1 != line.find(key):
                all_test[8][1][key].append(str(line[line.find("spends:") + 7:]))
    if -1 != line.find("cpu 1/4 resize with"):
        for key in all_test[7][1]:
            if -1 != line.find(key):
                all_test[7][1][key].append(str(line[line.find("spends:") + 7:]))

    if -1 != line.find("gpu to cpu with"):
        for key in all_test[14][1]:
            if -1 != line.find(key):
                all_test[14][1][key].append(str(line[line.find("spends:") + 7:]))
    if -1 != line.find("cpu to gpu with"):
        for key in all_test[13][1]:
            if -1 != line.find(key):
                all_test[13][1][key].append(str(line[line.find("spends:") + 7:]))

    # if -1 != line.find("\QA\Test_"):
    #     val = line[line.find("\QA\Test_") + 9: line.find("P.mp4")]
    #     for key in all_test[0][1]:
    #         if -1 != key.find(val):
    #             setvalinkey = key
    #             if times==5:
    #                 setTestItem += 1
    #             # all_test[0][1][key].append(str(XX))
    # if setvalinkey and -1 != line.find("perframetime ="):
    #     all_test[setTestItem][1][setvalinkey].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
    #     setvalinkey = ""
    #     times+=1
newf.close()

# cpu 版解码耗时
file = open('tmp.txt', 'w')
for testfil in  list:
    result = subprocess.check_output(["Test_Decode_9.exe",testfil,"MPDEC","0"])                 #['QA\\Test_1080P.mp4']
    print >> file, result
    print result
    haserr = result_has_error(result)
    if haserr:
        for key in all_test[0][1]:
            my = testfil[testfil.find("Test_")+5:testfil.find("P.mp4")]
            if -1 != key.find(my):
                all_test[0][1][key].append("FAIL")
                break

file.close()
newf = open("tmp.txt")
setvalinkey=""
for line in newf:
    # if -1 != line.find("1 FAILED TEST"):
    #     for ab in all_test[0][1]:
    #         all_test[0][1][ab].extends('fail')
    if -1 != line.find("QA\Test_"):
        val = line[line.find("QA\Test_") + 8: line.find("P.mp4")]
        for key in all_test[0][1]:
            if -1 != key.find(val):
                setvalinkey = key
                break
    if setvalinkey and -1 != line.find("decode perframetime = "):
        if not all_test[0][1][setvalinkey]:
            all_test[0][1][setvalinkey].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey = ""
newf.close()


# gpu 版dx9解码耗时
file = open('tmp.txt', 'w')
for testfil in  list:
    result = subprocess.check_output(["Test_Decode_9.exe",testfil,"SYSTEMDEC","1"])                 #['QA\\Test_1080P.mp4']
    print >> file, result
    print result
    haserr = result_has_error(result)
    if haserr:
        for key in all_test[1][1]:
            my = testfil[testfil.find("Test_")+5:testfil.find("P.mp4")]
            if -1 != key.find(my):
                all_test[1][1][key].append("FAIL")
                all_test[2][1][key].append("FAIL")
                break
file.close()

newf = open("tmp.txt")
setvalinkey=""
setvalinkey1=""
for line in newf:
    # if -1 != line.find("1 FAILED TEST"):
    #     for ab in all_test[0][1]:
    #         all_test[0][1][ab].extends('fail')
    if -1 != line.find("QA\Test_"):
        val = line[line.find("QA\Test_") + 8: line.find("P.mp4")]
        for key in all_test[1][1]:
            if -1 != key.find(val):
                setvalinkey = key
                setvalinkey1 = key
                break
    if setvalinkey and -1 != line.find("decode perframetime = "):
        if not all_test[1][1][setvalinkey]:
            all_test[1][1][setvalinkey].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey = ""

    if setvalinkey1 and -1 != line.find("End , perframetime ="):
        if not all_test[2][1][setvalinkey1]:
            all_test[2][1][setvalinkey1].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey1 = ""
newf.close()
########################################################################################################################################
# gpu 版dx10解码耗时
file = open('tmp.txt', 'w')
for testfil in  list:
    result = subprocess.check_output(["Test_Decode_10.exe",testfil,"SYSTEMDEC","1"])                 #['QA\\Test_1080P.mp4']
    print >> file, result
    print result
    haserr = result_has_error(result)
    if haserr:
        for key in all_test[3][1]:
            my = testfil[testfil.find("Test_")+5:testfil.find("P.mp4")]
            if -1 != key.find(my):
                all_test[3][1][key].append("FAIL")
                all_test[4][1][key].append("FAIL")
                break

file.close()
newf = open("tmp.txt")
setvalinkey=""
setvalinkey1=""
for line in newf:
    # if -1 != line.find("1 FAILED TEST"):
    #     for ab in all_test[0][1]:
    #         all_test[0][1][ab].extends('fail')
    if -1 != line.find("QA\Test_"):
        val = line[line.find("QA\Test_") + 8: line.find("P.mp4")]
        for key in all_test[3][1]:
            if -1 != key.find(val):
                setvalinkey = key
                setvalinkey1 = key
                break
    if setvalinkey and -1 != line.find("decode perframetime = "):
        if not all_test[3][1][setvalinkey]:
            all_test[3][1][setvalinkey].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey = ""
    if setvalinkey1 and -1 != line.find("End , perframetime ="):
        if not all_test[4][1][setvalinkey1]:
            all_test[4][1][setvalinkey1].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey1 = ""
newf.close()
################################################################################################################################################
# gpu 版dx11解码耗时
file = open('tmp.txt', 'w')
for testfil in  list:
    result = subprocess.check_output(["Test_Decode_11.exe",testfil,"SYSTEMDEC","1"])                 #['QA\\Test_1080P.mp4']
    print >> file, result
    print result
    haserr = result_has_error(result)
    if haserr:
        for key in all_test[5][1]:
            my = testfil[testfil.find("Test_")+5:testfil.find("P.mp4")]
            if -1 != key.find(my):
                all_test[5][1][key].append("FAIL")
                all_test[6][1][key].append("FAIL")
                break
file.close()

newf = open("tmp.txt")
setvalinkey=""
setvalinkey1=""
for line in newf:
    # if -1 != line.find("1 FAILED TEST"):
    #     for ab in all_test[0][1]:
    #         all_test[0][1][ab].extends('fail')
    if -1 != line.find("QA\Test_"):
        val = line[line.find("QA\Test_") + 8: line.find("P.mp4")]
        for key in all_test[5][1]:
            if -1 != key.find(val):
                setvalinkey = key
                setvalinkey1 = key
                break
    if setvalinkey and -1 != line.find("decode perframetime = "):
        if not all_test[5][1][setvalinkey]:
            all_test[5][1][setvalinkey].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey = ""
    if setvalinkey1 and -1 != line.find("End , perframetime ="):
        if not all_test[6][1][setvalinkey1]:
            all_test[6][1][setvalinkey1].append(str(line[line.find("perframetime =")+14:line.find(", fps =")]))
        setvalinkey1 = ""
newf.close()



ws["B10"]= "3840*2160"
ws["C10"]= "2560*1440"
ws["D10"]= "1920*1080"
ws["E10"]= "1280*720"
ws["F10"]= "960*540"

index = 11

for tests in all_test:
    colA = "A" + str(index)
    colB = "B" + str(index)
    colC = "C" + str(index)
    colD = "D" + str(index)
    colE = "E" + str(index)
    colF = "F" + str(index)
    ws[colA] = tests[0]

    ws[colB] = "".join(tests[1]["3840*2160"])
    ws[colC] = "".join(tests[1]["2560*1440"])
    ws[colD] = "".join(tests[1]["1920*1080"])
    ws[colE] = "".join(tests[1]["1280*720"])
    ws[colF] = "".join(tests[1]["960*540"])
    index+=1

fullname = "T-" + get_mac_address() +".xlsx"
ws.title = hostname
wb.save(fullname)



# for testfil in  list:
#     file = open('out.txt', 'w+')
#     result = subprocess.check_output(["Test_Decode_9.exe",testfil,"MPDEC","0"])
#     result = subprocess.check_output(["Test_Decode_9.exe", testfil, "SYSDEC", "1"])
#     result = subprocess.check_output(["Test_Decode_10.exe", testfil, "SYSDEC", "1"])
#     result = subprocess.check_output(["Test_Decode_11.exe", testfil, "SYSDEC", "1"])
#     print >> file, result
#     print result
#     file.close()


#mProcess = subprocess.Popen("usedlls.exe",stdin=subprocess.DEVNULL, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,universal_newlines=True)
#si = subprocess.STARTUPINFO()
#si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
#my_pro = subprocess.call("usedlls.exe",stdin=subprocess.PIPE,stderr=subprocess.PIPE,startupinfo=si)
#(output, error) = my_pro.communicate()
#my_pro = subprocess.Popen(cmd,stdin=subprocess.PIPE,stderr=subprocess.PIPE,startupinfo=si)
#subprocess.Popen("usedlls.exe")

def send_mail(mess):
    # 第三方 SMTP 服务
    mail_host = "smtp.163.com"  # 设置服务器
    mail_user = "770834094@qq.com"  # 用户名
    mail_pass = "wondershare123"  # 口令

    sender = 'sherwin_crawler@163.com'
    receivers = ["770834094@qq.com"]  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱

    # message = MIMEText(mess, 'plain', 'utf-8')
    message = MIMEMultipart()
    # message['From'] = '770834094@qq.com'    #匿名中转站
    # message['To'] = '280800024@qq.com'      #接收邮件的人

    subject = '性能测试报告邮件From'+ hostname + "@" + get_mac_address()
    message['Subject'] = Header(subject, 'utf-8')

    # 邮件正文内容
    message.attach(MIMEText(mess, 'plain', 'utf-8'))
    # 构造附件1，传送当前目录下的 test.txt 文件
    att1 = MIMEText(open(fullname, 'rb').read(), 'base64', 'utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # 这里的filename可以任意写，写什么名字，邮件中显示什么名字
    att1["Content-Disposition"] = 'attachment; filename=' + fullname
    message.attach(att1)

    try:
        smtpObj = smtplib.SMTP_SSL(mail_host, 465)
        smtpObj.login(mail_user, mail_pass)
        smtpObj.sendmail(sender, receivers, message.as_string())
        print "邮件发送成功"
    except smtplib.SMTPException:
        print "Error: 无法发送邮件"




send_mail("这是接口性能测试报告，来自"+ hostname + "@" + get_mac_address())
